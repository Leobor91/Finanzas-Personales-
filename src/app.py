from flask import Flask, request, jsonify, session
import sys
from pathlib import Path
from flask import render_template, redirect
import os
import json

# When executed as a script (python src/app.py) the package context is not set.
# Add project root to sys.path so absolute imports like `src.*` work.
if __package__ is None:
    project_root = Path(__file__).resolve().parent.parent
    sys.path.insert(0, str(project_root))

from src.infrastructure.database.postgres_adapter import PostgresMovementRepository
from src.infrastructure.database.postgres_adapter import PostgresMovementRepository
from datetime import datetime
from src.core.services.movement_service import MovementService
from src.core.domain.exceptions import InvalidAmountError, InvalidDateFormatError, InvalidTypeError

app = Flask(__name__, template_folder=str(Path(__file__).resolve().parent / 'templates'), static_folder=str(Path(__file__).resolve().parent / 'static'))

# Session secret
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-change-me')

from functools import wraps
from flask import url_for


def login_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('ui_login'))
        return f(*args, **kwargs)
    return wrapped


# Prevent browsers from caching authenticated pages so back-button won't show protected content
@app.after_request
def add_no_cache_headers(response):
    try:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    except Exception:
        pass
    return response


# Configuration for selectable database file
CONFIG_PATH = None


def get_repo():
    # Enforce Postgres-only: require DATABASE_URL
    db_url = os.environ.get('DATABASE_URL')
    if not db_url:
        raise RuntimeError('DATABASE_URL not set; application is configured for Postgres-only mode')
    # normalize scheme sometimes provided as postgres://
    if db_url.startswith('postgres://'):
        db_url = db_url.replace('postgres://', 'postgresql://', 1)
    return PostgresMovementRepository(db_url)


@app.route("/movements", methods=["POST"])
def create_movement():
    data = request.get_json() or {}
    # require account to be provided
    if not data.get('account'):
        return jsonify({'error': 'account requerido'}), 400
    # require authentication
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401

    repo = get_repo()
    service = MovementService(repo)
    try:
        movement_id = service.create_movement(
            date=data.get("date"),
            type=data.get("type"),
            amount=data.get("amount"),
            category=data.get("category"),
            description=data.get("description"),
            currency=data.get("currency", 'COP'),
            fx_rate=data.get("fx_rate", None),
            account=data.get("account"),
            user_id=uid,
        )
        return jsonify({"id": movement_id}), 201
    except InvalidAmountError:
        return jsonify({"error": "El monto debe ser un valor numérico mayor a cero"}), 400
    except InvalidDateFormatError:
        return jsonify({"error": "Formato de fecha incorrecto. Use AAAA-MM-DD"}), 400
    except InvalidTypeError:
        return jsonify({"error": "Tipo debe ser 'Ingreso' o 'Gasto'"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass



@app.route("/movements", methods=["GET"])
def list_movements():
    # Support a single-day query via ?date=YYYY-MM-DD which maps to from==to
    date_param = request.args.get("date")
    if date_param:
        date_from = date_to = date_param
    else:
        date_from = request.args.get("from")
        date_to = request.args.get("to")
    category = request.args.get("category")
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401
    repo = get_repo()
    try:
        from src.core.services.query_service import MovementQueryService

        qs = MovementQueryService(repo)
        results = qs.find(date_from=date_from, date_to=date_to, category=category, user_id=uid)
        return jsonify(results), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route("/reports/balance", methods=["GET"])
def report_balance():
    month = request.args.get("month")  # MM
    year = request.args.get("year")    # YYYY
    if not month or not year:
        return jsonify({"error": "Parámetros 'month' y 'year' son requeridos (MM, YYYY)."}), 400
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401

    repo = get_repo()
    try:
        from src.core.services.report_service import ReportService

        rs = ReportService(repo)
        # include previous month's net and cumulative net for the year (filter by user)
        bal = rs.monthly_with_carryover(month=month, year=year, user_id=uid)
        return jsonify(bal), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route("/reports/categories", methods=["GET"])
def report_categories():
    month = request.args.get('month')
    year = request.args.get('year')
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401

    repo = get_repo()
    try:
        from src.core.services.report_service import ReportService

        rs = ReportService(repo)
        rows = rs.expenses_by_category(year=year, month=month, user_id=uid)
        return jsonify([{"category": r.category, "total": r.total} for r in rows]), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route("/reports/top-expenses", methods=["GET"])
def report_top_expenses():
    month = request.args.get("month")  # MM
    year = request.args.get("year")    # YYYY
    if not month or not year:
        return jsonify({"error": "Parámetros 'month' y 'year' son requeridos (MM, YYYY)."}), 400
    limit = int(request.args.get("limit", 5))
    category = request.args.get("category")
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401

    repo = get_repo()
    try:
        from src.core.services.report_service import ReportService

        rs = ReportService(repo)
        rows = rs.top_expenses(month=month, year=year, limit=limit, category=category, user_id=uid)
        return jsonify(rows), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass

@app.route('/reports/years', methods=['GET'])
def report_years():
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401

    repo = get_repo()
    try:
        cur = repo.conn.cursor()
        cur.execute("SELECT DISTINCT to_char(date,'YYYY') as y FROM movements WHERE user_id = %s ORDER BY y DESC", (uid,))
        rows = cur.fetchall()
        years = [r[0] for r in rows if r[0]]
        return jsonify(years), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/reports/yearly', methods=['GET'])
def report_yearly():
    year = request.args.get('year')
    if not year:
        return jsonify({'error': "Parámetro 'year' requerido (YYYY)."}), 400
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401

    repo = get_repo()
    try:
        from src.core.services.report_service import ReportService
        rs = ReportService(repo)
        summary = rs.yearly_summary(year, user_id=uid)
        return jsonify(summary), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/reports/daily', methods=['GET'])
def report_daily():
    month = request.args.get('month')
    year = request.args.get('year')
    if not month or not year:
        return jsonify({'error': "Parámetros 'month' (MM) y 'year' (YYYY) son requeridos."}), 400
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401

    repo = get_repo()
    try:
        rows = repo.get_daily_aggregates(month, year, user_id=uid)
        return jsonify(rows), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/reports/export', methods=['GET'])
def report_export():
    # export movements between two dates as CSV: ?from=YYYY-MM-DD&to=YYYY-MM-DD
    date_from = request.args.get('from')
    date_to = request.args.get('to')
    category = request.args.get('category')
    mtype = request.args.get('type')
    if not date_from or not date_to:
        return jsonify({'error': "Parámetros 'from' y 'to' requeridos (YYYY-MM-DD)."}), 400
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401

    repo = get_repo()
    try:
        rows = repo.find_by_criteria(date_from=date_from, date_to=date_to, category=category, type_=mtype, user_id=uid)
        fmt = request.args.get('format', '').lower()
        # helper to parse amounts into numeric values
        def _to_number(val):
            if val is None:
                return 0.0
            try:
                return float(val)
            except Exception:
                s = str(val).strip()
                s = s.replace(' ', '')
                # Handle common thousands/decimal formats: '1.234.567,89' or '1,234,567.89'
                if s.count('.') > 0 and s.count(',') > 0:
                    # assume '.' thousands and ',' decimal
                    if s.rfind(',') > s.rfind('.'):
                        s = s.replace('.', '').replace(',', '.')
                else:
                    s = s.replace(',', '.')
                try:
                    return float(s)
                except Exception:
                    return 0.0

        total_amount = sum([_to_number(r.get('amount')) for r in rows])

        headers = ['id', 'date', 'type', 'account', 'category', 'description', 'amount', 'currency', 'fx_rate']
        # If the client requests Excel format, build an .xlsx file
        if fmt in ('xlsx', 'excel'):
            try:
                from openpyxl import Workbook
                from io import BytesIO
                wb = Workbook()
                ws = wb.active
                ws.append(headers)
                for r in rows:
                    amt = _to_number(r.get('amount'))
                    ws.append([
                        r.get('id'), r.get('date'), r.get('type'), r.get('account') or '', r.get('category') or '', r.get('description') or '', amt, r.get('currency') or 'COP', r.get('fx_rate') or ''
                    ])
                # blank row and total
                ws.append([''] * len(headers))
                ws.append(['', '', '', '', '', 'TOTAL', total_amount, '', ''])
                # set number format for amount column (G)
                for row in ws.iter_rows(min_row=2, min_col=7, max_col=7, max_row=ws.max_row):
                    for cell in row:
                        cell.number_format = '0.00'
                bio = BytesIO()
                wb.save(bio)
                bio.seek(0)
                from flask import Response
                suffix = f"_{mtype}" if mtype else ""
                filename = f"movements_{date_from}_to_{date_to}{suffix}.xlsx"
                resp = Response(bio.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
                return resp
            except Exception as e:
                return jsonify({'error': f'Error generando Excel: {str(e)}'}), 500

        # default: CSV
        import csv
        from io import StringIO
        si = StringIO()
        writer = csv.writer(si)
        writer.writerow(headers)
        def _to_number(val):
            if val is None:
                return 0.0
            try:
                return float(val)
            except Exception:
                s = str(val).strip()
                s = s.replace(' ', '')
                # Handle common thousands/decimal formats: '1.234.567,89' or '1,234,567.89'
                if s.count('.') > 0 and s.count(',') > 0:
                    # assume '.' thousands and ',' decimal
                    if s.rfind(',') > s.rfind('.'):
                        s = s.replace('.', '').replace(',', '.')
                else:
                    s = s.replace(',', '.')
                try:
                    return float(s)
                except Exception:
                    return 0.0

        total_amount = 0.0
        for r in rows:
            amt = _to_number(r.get('amount'))
            total_amount += amt
            writer.writerow([
                r.get('id'), r.get('date'), r.get('type'), r.get('account') or '', r.get('category') or '', r.get('description') or '', amt, r.get('currency') or 'COP', r.get('fx_rate') or ''
            ])
        # blank line for readability
        writer.writerow([])
        # place 'TOTAL' under description and the numeric summed amount under amount column
        writer.writerow(['', '', '', '', '', 'TOTAL', total_amount, '', ''])
        output = si.getvalue()
        from flask import Response
        suffix = f"_{mtype}" if mtype else ""
        filename = f"movements_{date_from}_to_{date_to}{suffix}.csv"
        resp = Response(output, mimetype='text/csv')
        resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/fx/latest', methods=['GET'])
def fx_latest():
    # Return latest exchange rates for COP to USD and EUR using exchangerate.host
    import urllib.request, json
    base = request.args.get('base', 'COP')
    symbols = request.args.get('symbols', 'USD,EUR')
    # First try exchangerate.host
    url1 = f'https://api.exchangerate.host/latest?base={base}&symbols={symbols}'
    try:
        with urllib.request.urlopen(url1, timeout=5) as resp:
            data = json.load(resp)
        # exchangerate.host may return {'success': False, 'error': {...}}
        if data.get('success', True) and 'rates' in data:
            return jsonify({'base': data.get('base', base), 'date': data.get('date'), 'rates': data.get('rates', {})}), 200
    except Exception:
        data = None

    # Fallback to open.er-api.com
    try:
        url2 = f'https://open.er-api.com/v6/latest/{base}'
        with urllib.request.urlopen(url2, timeout=5) as resp:
            data2 = json.load(resp)
        # data2 example: {'result':'success', 'rates': {'USD':0.00026, 'EUR':0.00024}, ...}
        rates2 = data2.get('rates', {})
        if rates2:
            # filter requested symbols
            wanted = {}
            for s in symbols.split(','):
                if s in rates2:
                    wanted[s] = rates2[s]
            return jsonify({'base': base, 'date': data2.get('time_last_update_utc') or data2.get('time_last_update_iso'), 'rates': wanted}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/categories', methods=['GET'])
def get_categories():
    type_q = request.args.get('type')
    if type_q not in ('Ingreso', 'Gasto'):
        return jsonify({'error': "Parámetro 'type' requerido y debe ser 'Ingreso' o 'Gasto'"}), 400
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401
    repo = get_repo()
    try:
        cats = repo.get_categories_by_type(type_q, user_id=uid)
        return jsonify(cats), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/categories', methods=['POST'])
def post_category():
    data = request.get_json() or {}
    type_q = data.get('type')
    name = data.get('name')
    icon = data.get('icon')
    if type_q not in ('Ingreso', 'Gasto') or not name:
        return jsonify({'error': "JSON debe contener 'type' ('Ingreso'|'Gasto') y 'name'"}), 400
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401
    repo = get_repo()
    try:
        cid = repo.add_category(type_q, name, icon, user_id=uid)
        return jsonify({'id': cid, 'name': name, 'icon': icon}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/categories/all', methods=['GET'])
def get_all_categories():
    # If the client prefers HTML, redirect to the UI page so the browser shows the cards
    # Avoid redirecting AJAX/fetch requests (which often set X-Requested-With)
    is_ajax = request.headers.get('X-Requested-With','').lower() == 'xmlhttprequest'
    accept_hdr = request.headers.get('Accept','')
    if 'text/html' in accept_hdr and not is_ajax:
        return redirect('/ui/categories')

    uid = session.get('user_id')
    repo = get_repo()
    try:
        cats = repo.list_all_categories(user_id=uid)
        return jsonify(cats), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/accounts', methods=['GET'])
def get_accounts():
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401
    repo = get_repo()
    try:
        # Return only accounts belonging to the authenticated user
        return jsonify(repo.get_accounts_with_balances(user_id=uid)), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/accounts/balances', methods=['GET'])
def get_accounts_balances():
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401
    repo = get_repo()
    try:
        # Ensure balances are returned only for the current user
        return jsonify(repo.get_accounts_with_balances(user_id=uid)), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/accounts', methods=['POST'])
def post_account():
    data = request.get_json() or {}
    name = data.get('name')
    initial = float(data.get('initial_balance', 0) or 0)
    currency = data.get('currency', 'COP')
    if not name:
        return jsonify({'error': 'name requerido'}), 400
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401
    repo = get_repo()
    try:
        aid = repo.add_account(name, initial, currency, user_id=uid)
        return jsonify({'id': aid, 'name': name, 'initial_balance': initial, 'currency': currency}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/ui/accounts')
@login_required
def ui_accounts():
    repo = get_repo()
    try:
        uid = session.get('user_id')
        accounts = repo.get_accounts_with_balances(user_id=uid)
        total = sum(a['balance'] for a in accounts)
        return render_template('accounts.html', accounts=accounts, total=total, active='accounts')
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/accounts/<account_name>/movements', methods=['GET'])
def get_account_movements(account_name):
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401
    repo = get_repo()
    try:
        movements = repo.get_movements_by_account(account_name, user_id=uid)
        balances = repo.get_accounts_with_balances()
        bal = next((a for a in balances if a['name'] == account_name), {'balance': 0.0, 'currency': 'COP'})
        return jsonify({'account': account_name, 'balance': bal.get('balance', 0.0), 'currency': bal.get('currency','COP'), 'movements': movements}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/movements/<int:mov_id>', methods=['DELETE'])
def delete_movement(mov_id):
    repo = get_repo()
    try:
        ok = repo.delete_movement(mov_id)
        if ok:
            return jsonify({'deleted': True}), 200
        return jsonify({'deleted': False}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/movements/<int:mov_id>', methods=['PUT'])
def put_movement(mov_id):
    data = request.get_json() or {}
    # sanitize fields
    allowed = ['date', 'type', 'amount', 'currency', 'fx_rate', 'category', 'description', 'account']
    payload = {k: data.get(k) for k in allowed if k in data}
    if not payload:
        return jsonify({'error': 'no fields to update'}), 400
    repo = get_repo()
    try:
        ok = repo.update_movement(mov_id, **payload)
        if ok:
            return jsonify({'updated': True}), 200
        return jsonify({'updated': False}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/accounts/transfer', methods=['POST'])
def post_account_transfer():
    data = request.get_json() or {}
    src_ac = data.get('from') or data.get('from_account')
    dst_ac = data.get('to') or data.get('to_account')
    try:
        amount = float(data.get('amount') or 0)
    except Exception:
        return jsonify({'error': 'amount inválido'}), 400
    description = data.get('description', 'Transferencia')
    date = data.get('date') or datetime.utcnow().strftime('%Y-%m-%d')
    if not src_ac or not dst_ac:
        return jsonify({'error': 'from y to son requeridos'}), 400
    if src_ac == dst_ac:
        return jsonify({'error': 'from y to deben ser cuentas diferentes'}), 400
    if amount <= 0:
        return jsonify({'error': 'amount debe ser mayor a cero'}), 400

    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401
    repo = get_repo()
    try:
        tid = repo.transfer_funds(src_ac, dst_ac, amount, date, description, user_id=uid)
        return jsonify({'transferred': True, 'transfer_id': tid}), 201
    except ValueError as ve:
        # insufficient funds
        if str(ve) == 'insufficient_funds':
            return jsonify({'error': 'Fondos insuficientes en la cuenta origen'}), 400
        return jsonify({'error': str(ve)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/transfers', methods=['GET'])
def get_transfers():
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401
    repo = get_repo()
    try:
        return jsonify(repo.list_transfers()), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/accounts/<account_name>/transfers', methods=['GET'])
def get_account_transfers(account_name):
    repo = get_repo()
    try:
        rows = repo.get_transfers_by_account(account_name)
        return jsonify(rows), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/transfers/<int:tid>', methods=['DELETE'])
def delete_transfer(tid):
    repo = get_repo()
    try:
        ok = repo.delete_transfer(tid)
        if ok:
            return jsonify({'deleted': True}), 200
        return jsonify({'deleted': False}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/transfers/<int:tid>', methods=['PUT'])
def put_transfer(tid):
    data = request.get_json() or {}
    # map client fields to DB columns
    allowed_map = {'date': 'date', 'from': 'from_account', 'to': 'to_account', 'amount': 'amount', 'currency': 'currency', 'description': 'description'}
    payload = {}
    for k, col in allowed_map.items():
        if k in data:
            payload[col] = data.get(k)
    if not payload:
        return jsonify({'error': 'no fields to update'}), 400
    repo = get_repo()
    try:
        ok = repo.update_transfer(tid, **payload)
        if ok:
            return jsonify({'updated': True}), 200
        return jsonify({'updated': False}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/movements/history', methods=['GET'])
def movements_history():
    # supports ?page=1&per_page=20 and optional date/category filters
    try:
        page = int(request.args.get('page', '1') or '1')
        per_page = int(request.args.get('per_page', '20') or '20')
    except Exception:
        return jsonify({'error': 'page and per_page must be integers'}), 400
    if page < 1 or per_page < 1:
        return jsonify({'error': 'page and per_page must be >= 1'}), 400

    date_from = request.args.get('from')
    date_to = request.args.get('to')
    category = request.args.get('category')
    mv_type = request.args.get('type')

    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401
    repo = get_repo()
    try:
        cur = repo.conn.cursor()
        # Use appropriate parameter placeholder depending on DB backend
        ph = '%s' if os.environ.get('DATABASE_URL') else '?'
        where = "WHERE 1=1"
        # restrict to user's rows
        where += " AND user_id = %s"
        params = [uid]
        if date_from:
            where += f" AND date >= {ph}"
            params.append(date_from)
        if date_to:
            where += f" AND date <= {ph}"
            params.append(date_to)
        if category:
            where += f" AND category LIKE {ph}"
            params.append(f"%{category}%")
        if mv_type:
            # accept exact type values (Ingreso/Gasto)
            where += f" AND type = {ph}"
            params.append(mv_type)

        # total count
        cnt_sql = f"SELECT COUNT(*) FROM movements {where}"
        cur.execute(cnt_sql, tuple(params))
        total = cur.fetchone()[0]

        # fetch paged rows
        offset = (page - 1) * per_page
        data_sql = f"SELECT id, to_char(date,'YYYY-MM-DD') as date, type, amount, currency, fx_rate, category, description, account FROM movements {where} ORDER BY date DESC, id DESC LIMIT {ph} OFFSET {ph}"
        exec_params = list(params) + [per_page, offset]
        cur.execute(data_sql, tuple(exec_params))
        rows = cur.fetchall()
        items = []
        for r in rows:
            items.append({
                'id': r[0], 'date': r[1], 'type': r[2], 'amount': r[3], 'currency': r[4], 'fx_rate': r[5], 'category': r[6], 'description': r[7], 'account': r[8]
            })

        total_pages = (total + per_page - 1) // per_page if per_page else 1
        return jsonify({'items': items, 'page': page, 'per_page': per_page, 'total': total, 'total_pages': total_pages}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/categories/<int:cat_id>', methods=['DELETE'])
def delete_category(cat_id):
    repo = get_repo()
    try:
        ok = repo.delete_category(cat_id)
        if ok:
            return jsonify({'deleted': True}), 200
        return jsonify({'deleted': False}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/categories/<int:cat_id>', methods=['PUT'])
def put_category(cat_id):
    data = request.get_json() or {}
    new_name = data.get('name')
    icon = data.get('icon')
    if not new_name:
        return jsonify({'error': "'name' requerido"}), 400
    repo = get_repo()
    try:
        ok = repo.update_category(cat_id, new_name)
        # If icon provided, update separately
        if ok and icon is not None:
            cur = repo.conn.cursor()
            cur.execute("UPDATE categories SET icon = %s WHERE id = %s", (icon, cat_id))
            try:
                repo.conn.commit()
            except Exception:
                pass
        if ok:
            return jsonify({'updated': True}), 200
        return jsonify({'updated': False}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


from flask import render_template


@app.route("/")
def ui_index():
    # If not authenticated, show login (home for anon users)
    if not session.get('user_id'):
        return redirect(url_for('ui_login'))
    return render_template('index.html', active='index')


@app.route("/ui/reports")
@login_required
def ui_reports():
    return render_template('reports.html', active='reports')


@app.route('/ui/categories')
@login_required
def ui_categories():
    repo = get_repo()
    try:
        uid = session.get('user_id')
        cats = repo.list_all_categories(user_id=uid)
        return render_template('categories.html', initial_categories=cats, active='categories')
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/ui/settings')
@login_required
def ui_settings():
    # list available .db files in cwd and data/
    db_files = []
    # In Postgres-only mode we don't expose local .db selection
    current = os.environ.get('DATABASE_URL')
    return render_template('settings.html', db_files=[], current_db=current, active='settings')


@app.route('/settings', methods=['POST'])
def post_settings():
    # accept form fields: db_name (text)
    raw = (request.form.get('db_name') or request.form.get('existing_db') or '').strip()
    if not raw:
        return jsonify({'error': 'db_name requerido'}), 400

    # Ensure .db extension
    p = Path(raw)
    if p.suffix == '':
        p = Path(str(p) + '.db')

    # Prefer storing DBs under data/ when relative
    if not p.is_absolute():
        data_dir = Path.cwd() / 'data'
        try:
            data_dir.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
        db_path = data_dir / p.name
        stored_value = str(Path('data') / p.name)
    else:
        db_path = p
        stored_value = str(p)

    # Postgres-only mode: changing local sqlite DB not supported
    return jsonify({'error': 'Operación no soportada en modo Postgres-only'}), 400


@app.route('/ui/cash-count')
@login_required
def ui_cash_count():
    return render_template('cash_count.html', active='cash_count')


@app.route('/ui/register', methods=['GET', 'POST'])
def ui_register():
    repo = None
    if request.method == 'GET':
        return render_template('register.html', active='register', error=None, success=None)

    # POST
    username = (request.form.get('username') or '').strip()
    password = request.form.get('password') or ''
    role = (request.form.get('role') or 'user').strip() or 'user'
    if not username or not password:
        return render_template('register.html', active='register', error='Usuario y contraseña requeridos', success=None)
    try:
        repo = get_repo()
        uid = repo.add_user(username, password, role)
        if uid:
            return render_template('register.html', active='register', error=None, success=True)
        return render_template('register.html', active='register', error='No se pudo crear el usuario', success=None)
    except Exception as e:
        return render_template('register.html', active='register', error=str(e), success=None)
    finally:
        try:
            if repo:
                repo.close()
        except Exception:
            pass



@app.route('/ui/login', methods=['GET', 'POST'])
def ui_login():
    if request.method == 'GET':
        return render_template('login.html', active='login', error=None)
    username = (request.form.get('username') or '').strip()
    password = request.form.get('password') or ''
    # server-side required check
    if not username or not password:
        # if AJAX request, return JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'ok': False, 'error': 'Usuario y contraseña requeridos'}), 400
        return render_template('login.html', active='login', error='Usuario y contraseña requeridos')
    repo = None
    try:
        repo = get_repo()
        uid = repo.authenticate_user(username, password)
        if uid:
            session['user_id'] = uid
            session['username'] = username
            # AJAX callers expect JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'ok': True}), 200
            return redirect('/')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'ok': False, 'error': 'Usuario o contraseña inválidos'}), 401
        return render_template('login.html', active='login', error='Usuario o contraseña inválidos')
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'ok': False, 'error': str(e)}), 500
        return render_template('login.html', active='login', error=str(e))
    finally:
        try:
            if repo:
                repo.close()
        except Exception:
            pass


@app.route('/logout')
def logout():
    # Clear server session and remove session cookie, then redirect to login
    session.clear()
    redirect_url = url_for('ui_login') + '?logged_out=1'
    resp = redirect(redirect_url)
    try:
        # remove session cookie on client
        resp.set_cookie(app.session_cookie_name, '', expires=0)
    except Exception:
        pass
    return resp


@app.route('/denominations', methods=['GET'])
def list_denominations_api():
    repo = get_repo()
    try:
        denoms = repo.list_denominations()
        return jsonify(denoms), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


@app.route('/denominations', methods=['POST'])
def post_denomination_api():
    data = request.get_json() or {}
    try:
        val = data.get('value')
        if val is None:
            return jsonify({'error': "'value' requerido"}), 400
        value = float(val)
    except Exception:
        return jsonify({'error': "'value' debe ser numérico"}), 400
    label = data.get('label')
    uid = session.get('user_id')
    if not uid:
        return jsonify({'error': 'autenticación requerida'}), 401
    repo = get_repo()
    try:
        did = repo.add_denomination(value, label, user_id=uid)
        return jsonify({'id': did, 'value': value, 'label': label}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        try:
            repo.close()
        except Exception:
            pass


if __name__ == "__main__":
    app.run(debug=True)
